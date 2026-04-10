#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Auphonic Folder Processor
-------------------------
Scans a specific folder for MP3 files and processes them through Auphonic.
- If no MP3s are found, it checks for WAVs and converts them to MP3.
- Uploads MP3s to Auphonic
- Starts production using credentials from Excel
- Downloads processed files back to the same folder
"""

import subprocess
import sys

import importlib.util

# ========== AUTO-INSTALL & BOOTSTRAP ==========
def bootstrap():
    packages = {
        "watchdog": "watchdog",
        "colorama": "colorama",
        "psutil": "psutil",
        "requests": "requests",
        "openpyxl": "openpyxl",
        "requests_toolbelt": "requests-toolbelt",
        "bidi": "python-bidi"
    }
    import importlib.util
    for module, package in packages.items():
        if importlib.util.find_spec(module) is None:
            print(f"📦 Missing library detected: Installing {package}...")
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", package, "--quiet"])
            except Exception as e:
                print(f"⚠️ Failed to install {package}: {e}")

bootstrap()

import os
import time
import json
import shutil
import concurrent.futures
from pathlib import Path
from urllib.parse import urlparse, unquote
from datetime import timedelta

import colorama
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import openpyxl
from requests_toolbelt.multipart.encoder import MultipartEncoder, MultipartEncoderMonitor
from typing import List, Dict, Any, Optional, Tuple, Union

colorama.just_fix_windows_console()

# ======================================================
#                 CONFIG
# ======================================================
# Path to the settings Excel file
EXCEL_PATH = r"E:\HABASHY\Python Codes\settings.xlsx"

# Helper to find tool paths (Portable Mode)
def find_tool(name, fallback):
    # 1. Check local bin/ folder
    local_bin = Path(__file__).parent / "bin" / f"{name}.exe"
    if local_bin.exists(): return str(local_bin)
    # 2. Check system path
    system_path = shutil.which(name)
    if system_path: return system_path
    # 3. Fallback to hardcoded
    return fallback

# Path to FFMPEG and FFPROBE
BIN_FALLBACK = r"E:\HABASHY\ffmpeg-2025-10-21-git-535d4047d3-full_build\ffmpeg-2025-10-21-git-535d4047d3-full_build\bin"
FFMPEG_PATH = find_tool("ffmpeg", os.path.join(BIN_FALLBACK, "ffmpeg.exe"))
FFPROBE_PATH = find_tool("ffprobe", os.path.join(BIN_FALLBACK, "ffprobe.exe"))

# Auphonic credits calculation
MIN_PER_CREDIT = 60

# ======================================================
#                   COLORS + ICONS
# ======================================================
RESET  = "\033[0m"; BOLD   = "\033[1m"
RED   = "\033[31m"; GREEN = "\033[32m"; YELLOW = "\033[33m"
BLUE  = "\033[34m"; CYAN  = "\033[36m"; MAGENTA = "\033[35m"; BLACK = "\033[30m"
WHITE = "\033[37m"

ICON_START = f"{BLUE}>> {RESET}"
ICON_OK    = f"{GREEN}[OK] {RESET}"
ICON_FAIL  = f"{RED}[FAIL] {RESET}"
ICON_WARN  = f"{YELLOW}[WARN] {RESET}"
ICON_INFO  = f"{CYAN}[INFO] {RESET}"
ICON_TIME  = f"{CYAN}.. {RESET}"
ICON_STEP  = f"{MAGENTA}-> {RESET}"

PROGRESS_FULL = "="
PROGRESS_EMPTY = "-"

# ======================================================
#              HELPER FUNCTIONS
# ======================================================
_global_session = None

def get_session():
    """Returns a global, shared requests Session with retry logic."""
    global _global_session
    if _global_session is None:
        _global_session = requests.Session()
        retry = Retry(
            total=10,
            connect=10,
            read=10,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["HEAD", "GET", "PUT", "POST", "DELETE"]
        )
        # Increase connection pool sizes to handle multithreading comfortably
        adapter = HTTPAdapter(max_retries=retry, pool_connections=20, pool_maxsize=20)
        _global_session.mount("https://", adapter)
        _global_session.mount("http://", adapter)
    return _global_session

def fmt(sec):
    try:
        sec = int(sec)
    except:
        sec = 0
    return str(timedelta(seconds=sec))

def print_header(title):
    line = "=" * 60
    print(f"\n{BLUE}{line}{RESET}")
    print(f"{ICON_START} {BOLD}{WHITE}{title}{RESET}")
    print(f"{BLUE}{line}{RESET}")

def warn(msg):
    print(f"{ICON_WARN} {YELLOW}{msg}{RESET}")

def fail(msg):
    print(f"{ICON_FAIL} {RED}{msg}{RESET}")

def ok(msg):
    print(f"{ICON_OK} {GREEN}{msg}{RESET}")

def get_duration(path: Path) -> float:
    try:
        r = subprocess.run(
            [FFPROBE_PATH, "-v", "error", "-show_entries", "format=duration",
             "-of", "default=noprint_wrappers=1:nokey=1", str(path)],
            capture_output=True, text=True
        )
        return float(r.stdout.strip())
    except Exception as e:
        warn(f"Could not get duration for {path.name}: {e}. Ensure ffprobe.exe exists at {FFPROBE_PATH}")
        return 0.0

def convert_wav_to_mp3(wav_path, mp3_path):
    """Converts a WAV file to MP3 using ffmpeg."""
    cmd = [
        FFMPEG_PATH, "-y",
        "-i", str(wav_path),
        "-codec:a", "libmp3lame",
        "-b:a", "320k",
        str(mp3_path)
    ]
    try:
        print(f"{ICON_STEP} Converting {wav_path.name} to MP3...")
        subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        ok(f"Converted: {mp3_path.name}")
        return True
    except subprocess.CalledProcessError:
        fail(f"Failed to convert {wav_path.name}. Is ffmpeg installed?")
        return False
    except Exception as e:
        fail(f"Error converting {wav_path.name}: {e}")
        return False

# ======================================================
#              AUPHONIC LOGIC
# ======================================================
def load_accounts():
    if not os.path.isfile(EXCEL_PATH):
        fail(f"Excel not found: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb[wb.sheetnames[0]]

    accounts: List[Dict[str, Any]] = []
    
    # Use count to avoid the 'row += 1' type inference issue in some checkers
    import itertools
    for row in itertools.count(start=2):
        email = sheet[f"A{row}"].value
        api   = sheet[f"B{row}"].value
        preset= sheet[f"C{row}"].value

        if not (email and api and preset):
            break

        accounts.append({
            "email": str(email),
            "api": str(api),
            "preset": str(preset),
            "remaining_minutes": 0.0
        })

    if not accounts:
        fail("No accounts found in Excel.")
        sys.exit(1)

    return accounts

def save_accounts_to_excel(accounts: List[Dict[str, Any]]):
    """Overwrites the Excel file with the provided list of active accounts."""
    while True:
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            # Headers
            sheet["A1"] = "Email"
            sheet["B1"] = "API Key"
            sheet["C1"] = "Preset"
            
            for i, acc in enumerate(accounts, start=2):
                sheet[f"A{i}"] = acc["email"]
                sheet[f"B{i}"] = acc["api"]
                sheet[f"C{i}"] = acc["preset"]
                
            wb.save(EXCEL_PATH)
            ok(f"Updated {EXCEL_PATH} (Active accounts: {len(accounts)})")
            break # Success
        except PermissionError:
            fail(f"Permission Denied: Could not save to {EXCEL_PATH}")
            warn("Is the file open in Excel? Please CLOSE IT and press Enter to retry, or Ctrl+C to skip.")
            input("Press Enter to retry...")
        except Exception as e:
            fail(f"Could not update Excel: {e}")
            break

def cleanup_empty_accounts(accounts: List[Dict[str, Any]]):
    """Checks for accounts with 0 balance and removes them from Excel."""
    to_keep = [acc for acc in accounts if float(acc.get("remaining_minutes", 0)) > 0.1]
    removed_count = len(accounts) - len(to_keep)
    
    if removed_count > 0:
        warn(f"Detected {removed_count} account(s) with zero balance. Removing from Excel...")
        save_accounts_to_excel(to_keep)
        return to_keep
    return accounts

def get_account_remaining_minutes(acc: Dict[str, Any]) -> float:
    url = "https://auphonic.com/api/user.json"
    headers_dict = {"Authorization": f"Bearer {acc.get('api', '')}"}
    try:
        r = get_session().get(url, headers=headers_dict, timeout=10)
        r.raise_for_status()
        data = r.json().get("data", {})
        credits_val = float(data.get("credits", 0))
        return credits_val * MIN_PER_CREDIT
    except Exception as e:
        warn(f"Error fetching credits for {acc.get('email', 'unknown')}: {e}")
        return 0.0

def headers(acc: Dict[str, Any]) -> Dict[str, str]:
    return {"Authorization": f"Bearer {acc.get('api', '')}"}

def auphonic_create(input_mp3: str, acc: Dict[str, Any], show_progress: bool = True) -> str:
    url = "https://auphonic.com/api/simple/productions.json"
    file_size = os.path.getsize(input_mp3)

    def _make_monitor():
        fields = {
            "preset": acc["preset"],
            "title": os.path.basename(input_mp3),
            "action": "start",
            "bitrate": "320",
            "input_file": (os.path.basename(input_mp3), open(input_mp3, "rb"))
        }
        print(f"{ICON_INFO} Using Preset: {acc['preset']}")
        encoder = MultipartEncoder(fields=fields)
        start_time = time.time()

        def _cb(monitor):
            pct = monitor.bytes_read / file_size if file_size else 0
            pct = max(0.0, min(1.0, pct))
            elapsed = max(time.time() - start_time, 1e-6)
            speed = monitor.bytes_read / elapsed
            bar_len = 30
            filled = int(bar_len * pct)
            bar = PROGRESS_FULL * filled + PROGRESS_EMPTY * (bar_len - filled)
            sys.stdout.write(
                f"\r{ICON_STEP} Uploading {os.path.basename(input_mp3)} "
                f"[{MAGENTA}{bar}{RESET}] {int(pct*100):3d}%"
            )
            sys.stdout.flush()

        return MultipartEncoderMonitor(encoder, _cb)

    if show_progress:
        monitor = _make_monitor()
        headers_upload = headers(acc)
        headers_upload["Content-Type"] = monitor.content_type
        data = monitor
    else:
        m = MultipartEncoder(fields={
            "preset": acc["preset"],
            "title": os.path.basename(input_mp3),
            "action": "start",
            "bitrate": "320",
            "input_file": (os.path.basename(input_mp3), open(input_mp3, "rb"))
        })
        print(f"{ICON_INFO} Using Preset: {acc['preset']}")
        headers_upload = headers(acc)
        headers_upload["Content-Type"] = m.content_type
        data = m

    r = get_session().post(url, headers=headers_upload, data=data)
    if show_progress: print()
    r.raise_for_status()
    uuid = r.json()["data"]["uuid"]
    return uuid

def auphonic_wait(uuid: str, acc: Dict[str, Any], show_progress: bool = True) -> bool:
    url = f"https://auphonic.com/api/production/{uuid}/status.json"
    
    while True:
        try:
            r = get_session().get(url, headers=headers(acc))
            r.raise_for_status()
            data = r.json()["data"]
            status = data["status"]
            msg = data["status_string"] or ""
            
            if show_progress:
                sys.stdout.write(f"\r{ICON_TIME} Status: {CYAN}{msg}{RESET}   ")
                sys.stdout.flush()

            if status == 3: # Done
                if show_progress: print()
                return True
            if status == 2: # Error
                if show_progress: print()
                err_msg = data.get("error_message") or data.get("error_status") or msg
                fail(f"Auphonic error: {err_msg}")
                return False
        except Exception as e:
            if show_progress:
                sys.stdout.write(f"\r{ICON_WARN} {YELLOW}Network delay, retrying...{RESET}          ")
                sys.stdout.flush()
                
        time.sleep(3)
        
    return False

def download_outputs(uuid: str, acc: Dict[str, Any], dest_dir: Path, show_progress: bool = True) -> None:
    url = f"https://auphonic.com/api/production/{uuid}.json"
    try:
        r = get_session().get(url, headers=headers(acc))
        r.raise_for_status()
    except Exception as e:
        fail(f"Fetch production failed: {e}")
        return

    data = r.json().get("data", {})
    outs = data.get("output_files", []) or []
    
    for out in outs:
        dl = out.get("download_url") or ""
        ending = out.get("ending") or ""
        
        if dl.startswith("/"): dl = "https://auphonic.com" + dl
        if not dl: continue

        # Determine filename
        filename = "output" + (f".{ending}" if ending else "")
        try:
            parsed = urlparse(dl)
            name = Path(parsed.path).name
            if name: filename = unquote(name.split("?")[0])
        except: pass

        target = dest_dir / filename
        
        try:
            with get_session().get(dl, headers=headers(acc), stream=True) as resp:
                resp.raise_for_status()
                total = int(resp.headers.get("content-length") or 0)
                downloaded = 0
                
                with open(target, "wb") as f:
                    for chunk in resp.iter_content(8192):
                        if not chunk: continue
                        f.write(chunk)
                        downloaded += len(chunk)
                        
                        if show_progress and total > 0:
                            pct = downloaded / total
                            bar_len = 30
                            filled = int(bar_len * pct)
                            bar = PROGRESS_FULL * filled + PROGRESS_EMPTY * (bar_len - filled)
                            sys.stdout.write(
                                f"\r{ICON_STEP} Downloading {filename} "
                                f"[{MAGENTA}{bar}{RESET}] {int(pct*100):3d}%"
                            )
                            sys.stdout.flush()
                if show_progress: print()
                ok(f"Downloaded: {filename}")
        except Exception as e:
            warn(f"Download failed: {e}")

# ======================================================
#              MAIN LOGIC
# ======================================================
def process_folder(folder_path):
    folder = Path(folder_path)
    if not folder.exists():
        fail(f"Folder not found: {folder}")
        return

    print_header(f"Scanning: {folder}")
    
    # 1. Find MP3s
    mp3s = sorted([f for f in folder.iterdir() if f.suffix.lower() == '.mp3'])
    
    if not mp3s:
        warn("No .mp3 files found. Checking for .wav files...")
        wavs = sorted([f for f in folder.iterdir() if f.suffix.lower() == '.wav'])
        if wavs:
            print(f"{ICON_INFO} Found {len(wavs)} WAV file(s). Converting to MP3...")
            for wav in wavs:
                mp3_target = wav.with_suffix('.mp3')
                if convert_wav_to_mp3(wav, mp3_target):
                    mp3s.append(mp3_target)
            
            # Re-sort to be safe and ensure processed matches
            mp3s.sort()
            print(f"{ICON_INFO} Conversion complete. Final list: {len(mp3s)} MP3 file(s)")
        else:
            warn("No .mp3 or .wav files found in this folder.")
            return
    else:
        print(f"{ICON_INFO} Found {len(mp3s)} MP3 file(s)")

    # 2. Load Accounts & Check Credits
    accounts = load_accounts()
    print(f"{ICON_INFO} Loaded {len(accounts)} account(s)")
    
    # Update credits
    print(f"{ICON_INFO} Checking account credits...")
    total_available_min: float = 0.0
    for acc in accounts:
        rem = float(get_account_remaining_minutes(acc))
        acc["remaining_minutes"] = rem
        total_available_min += rem
        email_str = str(acc.get("email", "Unknown"))
        print(f"   {CYAN}* {email_str}: {rem:.2f} minutes available{RESET}")
    
    # NEW: Automatic cleanup of zero-balance accounts
    accounts = cleanup_empty_accounts(accounts)
    
    print(f"{ICON_INFO} Total Available: {BOLD}{total_available_min:.2f} minutes{RESET}")

    # Sort accounts by credits DESCENDING so we use the fullest accounts first
    accounts.sort(key=lambda x: float(x.get("remaining_minutes", 0)), reverse=True)

    # 3. Plan Uploads
    uploads = []
    print(f"\n{ICON_INFO} {BOLD}Creating Processing Plan:{RESET}")
    for mp3 in mp3s:
        duration = get_duration(mp3)
        duration_min = max(duration / 60, 0.01)
        
        # Try to find an account with enough credits
        target_acc: Dict[str, Any] = {}
        found = False
        for acc in accounts:
            if float(acc.get("remaining_minutes", 0)) >= duration_min:
                target_acc = acc
                found = True
                break
        
        if found:
            current_rem = float(target_acc.get("remaining_minutes", 0.0))
            new_rem = current_rem - duration_min
            target_acc["remaining_minutes"] = new_rem
            
            uploads.append((mp3, target_acc))
            
            acc_email = str(target_acc.get("email", "Unknown"))
            print(f"   {ICON_STEP} {mp3.name:<25} ({CYAN}{duration_min:5.2f} min{RESET}) -> {acc_email} "
                  f"[{WHITE}Rem: {new_rem:.2f} min{RESET}]")
        else:
            fail(f"Skipping {mp3.name}: No account has {duration_min:.2f} min available.")

    if not uploads:
        warn("No uploads planned due to insufficient credits or no files.")
        return

    # Print Final Summary Table
    print(f"\n{BLUE}{'='*60}{RESET}")
    print(f"{BOLD}{WHITE}PLAN SUMMARY{RESET}")
    print(f"{BLUE}{'-'*60}{RESET}")
    for acc in accounts:
        # assigned type: List[Tuple[Path, Dict[str, Any]]]
        assigned = [u for u in uploads if u[1] == acc]
        if assigned:
            sub_total: float = sum(max(get_duration(u[0])/60, 0.01) for u in assigned)
            acc_email = str(acc.get("email", "Unknown"))
            print(f"{BOLD}{acc_email}{RESET}")
            print(f"  Files: {len(assigned)} | Total Task Time: {sub_total:.2f} min")
            rem_val = float(acc.get("remaining_minutes", 0.0))
            print(f"  Projected Account Balance: {GREEN}{rem_val:.2f} min{RESET}")
    print(f"{BLUE}{'='*60}{RESET}")

    # 4. Execute
    print_header("Starting Processing")
    
    def process_item(item: Tuple[Path, Dict[str, Any]]) -> bool:
        mp3_path, acc = item
        try:
            print(f"\n{ICON_START} Processing: {mp3_path.name}")
            
            # Upload
            acc_dict: Dict[str, Any] = acc
            uuid = auphonic_create(str(mp3_path), acc_dict)
            ok(f"Uploaded {mp3_path.name}")
            
            # Wait
            if auphonic_wait(uuid, acc_dict):
                # Download
                download_outputs(uuid, acc_dict, folder)
                return True
            return False
        except Exception as e:
            fail(f"Error processing {mp3_path.name}: {e}")
            return False

    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        list(executor.map(process_item, uploads))

    print_header("Done")
    ok("All files processed.")

if __name__ == "__main__":
    print(f"{BOLD}Auphonic Batch Processor{RESET}")
    print("Paste the folder path containing MP3 files:")
    
    try:
        folder_input = input("Folder Path > ").strip()
        # Remove quotes if user pasted path with quotes
        folder_input = folder_input.replace('"', '').replace("'", "")
        
        if folder_input:
            process_folder(folder_input)
        else:
            fail("No folder provided.")
    except KeyboardInterrupt:
        print("\nCancelled.")
    except Exception as e:
        fail(f"Unexpected error: {e}")
    

    input("\nPress Enter to exit...")

    


