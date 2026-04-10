import os
import sys
import time
import requests
import openpyxl
from pathlib import Path

# ======================================================
#                 CONFIGURATION
# ======================================================
# You can hardcode your API key here or paste it when running
DEFAULT_API_KEY = "FPSX8439695c891790fb28127bb7cfc9cbab" 

# Column name in Excel that contains the prompts
PROMPT_COLUMN_HEADER = "Prompt"

# Output folder for images
OUTPUT_FOLDER = "Generated_Images"

# ======================================================
#                 COLORS & UI
# ======================================================
RESET  = "\033[0m"
GREEN  = "\033[32m"
RED    = "\033[31m"
CYAN   = "\033[36m"
YELLOW = "\033[33m"

def print_info(msg): print(f"{CYAN}[INFO] {msg}{RESET}")
def print_success(msg): print(f"{GREEN}[OK] {msg}{RESET}")
def print_error(msg): print(f"{RED}[ERROR] {msg}{RESET}")
def print_warn(msg): print(f"{YELLOW}[WARN] {msg}{RESET}")

# ======================================================
#                 FREEPIK API
# ======================================================
def generate_image(prompt, api_key):
    """
    Sends a prompt to the Freepik API and returns the image content.
    Note: You might need to adjust the URL or parameters based on the specific 
    Freepik API documentation version you are using.
    """
    # URL for Freepik Text to Image
    # Using the Flux Dev model as it is a common standard endpoint
    url = "https://api.freepik.com/v1/ai/text-to-image/flux-dev" 
    
    headers = {
        "x-freepik-api-key": api_key, 
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    
    payload = {
        "prompt": prompt,
        "num_images": 1,
        "image_size": "square_hd" # Adjusted parameter name for Flux endpoint often differs, usually 'image_size' or 'size'
    }

    try:
        # Note: If the API expects a different structure, modify this request
        response = requests.post(url, json=payload, headers=headers)
        
        if response.status_code == 200:
            data = response.json()
            
            # Handle Async Response (Flux/Mystic models usually return a task_id)
            if "data" in data and "task_id" in data["data"]:
                task_id = data["data"]["task_id"]
                print_info(f"Task initiated. ID: {task_id}. Waiting for completion...")
                
                # Polling loop
                poll_url = f"https://api.freepik.com/v1/ai/text-to-image/flux-dev/{task_id}"
                # Alternatively try the general one if above fails: f"https://api.freepik.com/v1/ai/tasks/{task_id}"
                
                for attempt in range(30): # Wait up to 60 seconds (30 * 2s)
                    time.sleep(2)
                    poll_res = requests.get(poll_url, headers=headers)
                    if poll_res.status_code == 200:
                        poll_data = poll_res.json()
                        status = poll_data.get("data", {}).get("status", "")
                        
                        if status == "COMPLETED":
                             generated = poll_data.get("data", {}).get("generated", [])
                             if generated and len(generated) > 0:
                                 img_obj = generated[0]
                                 if "base64" in img_obj:
                                     import base64
                                     return base64.b64decode(img_obj["base64"])
                                 elif "url" in img_obj:
                                     return requests.get(img_obj["url"]).content
                        elif status == "FAILED":
                            print_error(f"Task failed: {poll_data}")
                            return None
                        # If PENDING or IN_PROGRESS, continue loop
                    else:
                        print_warn(f"Polling warning ({poll_res.status_code}): {poll_res.text}")
                
                print_error("Timeout waiting for image generation.")
                return None

            # Handle Sync Response (Legacy)
            elif "data" in data and isinstance(data["data"], list) and len(data["data"]) > 0:
                img_data = data["data"][0]
                if "base64" in img_data:
                    import base64
                    return base64.b64decode(img_data["base64"])
                elif "url" in img_data:
                    return requests.get(img_data["url"]).content
            
            print_error(f"Could not find image in response: {data}")
            return None
        else:
            print_error(f"API Error ({response.status_code}): {response.text}")
            return None
    except Exception as e:
        import traceback
        traceback.print_exc()
        print_error(f"Request failed: {e}")
        return None

# ======================================================
#                 MAIN LOGIC
# ======================================================
def process_excel(file_path, api_key):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active # Use the first sheet

    # Find the prompt column
    prompt_col_idx = None
    status_col_idx = None
    
    header_row = 1
    for cell in sheet[header_row]:
        if cell.value == PROMPT_COLUMN_HEADER:
            prompt_col_idx = cell.column
        if cell.value == "Status":
            status_col_idx = cell.column

    if prompt_col_idx is None:
        print_error(f"Column '{PROMPT_COLUMN_HEADER}' not found in Excel.")
        return

    # Create Status column if it doesn't exist
    if status_col_idx is None:
        status_col_idx = sheet.max_column + 1
        sheet.cell(row=header_row, column=status_col_idx).value = "Status"

    output_dir = Path(os.path.dirname(file_path)) / OUTPUT_FOLDER
    output_dir.mkdir(exist_ok=True)
    print_info(f"Saving images to: {output_dir}")

    rows = list(sheet.iter_rows(min_row=2))
    total = len(rows)
    
    print_info(f"Found {total} rows to process.")

    for i, row in enumerate(rows, start=1):
        cell_prompt = row[prompt_col_idx - 1]
        cell_status = row[status_col_idx - 1]
        
        prompt_text = cell_prompt.value
        
        if not prompt_text:
            continue
            
        if cell_status.value == "Done":
             print_info(f"Skipping row {i} (Already Done)")
             continue

        print(f"[{i}/{total}] Generating: {prompt_text[:30]}...")
        
        image_content = generate_image(prompt_text, api_key)
        
        if image_content:
            # Save Image
            safe_filename = "".join([c for c in prompt_text[:20] if c.isalnum() or c in (' ', '-', '_')]).strip()
            filename = f"{i}_{safe_filename}.png"
            file_path_img = output_dir / filename
            
            with open(file_path_img, "wb") as f:
                f.write(image_content)
            
            print_success(f"Saved: {filename}")
            cell_status.value = "Done"
        else:
            cell_status.value = "Failed"
        
        # Save progress every row (optional, safer)
        wb.save(file_path)
        time.sleep(1) # Be nice to the API

    print_success("Processing complete!")

if __name__ == "__main__":
    os.system('cls' if os.name == 'nt' else 'clear')
    print(f"{CYAN}=== Freepik Prompt Processor ==={RESET}")
    
    excel_path = input("Enter Excel file path: ").strip().strip('"')
    
    if not os.path.exists(excel_path):
        print_error("File not found.")
        sys.exit(1)
        
    api_key = DEFAULT_API_KEY
    if not api_key:
        api_key = input("Enter Freepik API Key: ").strip()
        
    if not api_key:
        print_error("API Key is required.")
        sys.exit(1)
        
    process_excel(excel_path, api_key)
