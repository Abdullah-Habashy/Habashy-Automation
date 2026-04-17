#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
habashy automation
------------------------------
Auto Project Media Importer + WAV Merge + Auphonic Pipeline

- Per-day project folder under BASE_ROOT -> dd.mm.yyyy
- Subfolders: audio / video / export / compressed
- Copies AUDIO first, then runs WAV->MP3->Auphonic pipeline inside /audio
- Then copies VIDEO
- JSON DB to prevent re-importing the same media from cards
- Progress bars for copying and FFmpeg
- Uses Excel settings for Auphonic accounts
- Chooses the account by remaining Auphonic credits
- Parallel Auphonic uploads/processing for speed
"""

import os
import sys
import ctypes
import string
import re
import time
import json
import shutil
import subprocess
import threading
import concurrent.futures
from pathlib import Path
from urllib.parse import urlparse, unquote
from datetime import datetime, timedelta

# --- Bootstrap: ensure third-party packages are installed before importing them ---
import importlib

THIRD_PARTY_MAP = {
    # module_name: pip_package_name
    "colorama": "colorama",
    "requests": "requests",
    "openpyxl": "openpyxl",
    "requests_toolbelt": "requests-toolbelt",
    "bidi": "python-bidi",
    "urllib3": "urllib3",
}

def ensure_packages(pack_map):
    missing = []
    for mod, pkg in pack_map.items():
        try:
            importlib.import_module(mod)
        except Exception:
            missing.append(pkg)

    if not missing:
        return True

    print(f"Installing missing packages: {', '.join(missing)}")
    for pkg in missing:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])
        except Exception as e:
            print(f"Failed to install {pkg}: {e}")
            return False
    return True


# Try to ensure required third-party packages are present (best-effort)
ensure_packages(THIRD_PARTY_MAP)

# Try to import bidi for Arabic RTL support (optional)
try:
    from bidi.algorithm import get_display
    HAS_BIDI = True
except Exception:
    HAS_BIDI = False
    def get_display(text):
        return text  # Fallback: return as-is


import colorama
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import openpyxl
from requests_toolbelt.multipart.encoder import MultipartEncoder, MultipartEncoderMonitor

colorama.just_fix_windows_console()

# ======================================================
#                 CONFIG — EDIT HERE
# ======================================================
BASE_ROOT   = Path(r"F:\2026\3RD").resolve()
# NOTE: These are now RELATIVE paths to search for on any drive
# e.g. X:/PRIVATE/M4ROOT/CLIP
VIDEO_SRC_REL = Path(r"PRIVATE\M4ROOT\CLIP")
AUDIO_SRC_REL = Path(r"STEREO\FOLDER01")

SUBFOLDERS   = ["audio", "video", "export", "compressed"]

# Helper to find tool paths (Portable Mode)
def find_tool(name, fallback):
    # 1. Check local bin/ folder
    local_bin = Path(__file__).parent / "bin" / f"{name}.exe"
    if local_bin.exists(): return str(local_bin)
    # 2. Check system path
    system_path = shutil.which(name)
    if system_path: return system_path
    # 3. Fallback to hardcoded
    return fallback

# Path to FFMPEG and FFPROBE
BIN_FALLBACK = r"E:\HABASHY\ffmpeg-2025-10-21-git-535d4047d3-full_build\ffmpeg-2025-10-21-git-535d4047d3-full_build\bin"
FFMPEG_PATH = find_tool("ffmpeg", os.path.join(BIN_FALLBACK, "ffmpeg.exe"))
FFPROBE_PATH = find_tool("ffprobe", os.path.join(BIN_FALLBACK, "ffprobe.exe"))

VIDEO_EXTS = {".mp4"}
AUDIO_EXTS = {".wav"}

# Minimum code allowed from filenames
MIN_VIDEO_CODE  = 5082     # C5071 and above
MIN_AUDIO_CODE  = 1254     # ZOOM1249 and above

# Persistent DB for imported media (prevents duplicate imports from cards)
COPIED_DB_PATH   = BASE_ROOT / ".copied_files_db.json"
FALLBACK_DB_PATH = Path(__file__).parent / ".copied_files_db_backup.json"

CLI_MODE = "minimal"  # "plain" (no colors), "minimal" (white with colored start/success), "full" (full colors/icons)
IS_PLAIN = CLI_MODE == "plain"
IS_MINIMAL = CLI_MODE == "minimal"
IS_FULL = CLI_MODE == "full"

CHUNK_SIZE = 4 * 1024 * 1024
COPY_BAR_WIDTH = 24 if IS_PLAIN or IS_MINIMAL else 28
PROGRESS_FULL = "#" if (IS_PLAIN or IS_MINIMAL) else "█"
PROGRESS_EMPTY = "." if (IS_PLAIN or IS_MINIMAL) else "░"
FFMPEG_BAR_LEN = 32 if (IS_PLAIN or IS_MINIMAL) else 40

# ======================================================
#                TELEGRAM NOTIFICATION
# ======================================================
TELEGRAM_BOT_TOKEN = ""  # Get from @BotFather

# يمكنك وضع Chat ID واحد أو قائمة من IDs
# مثال لـ ID واحد: TELEGRAM_CHAT_ID = "123456789"
# مثال لعدة IDs: TELEGRAM_CHAT_ID = ["123456789", "-1001234567890"]
TELEGRAM_CHAT_ID = "-5034031577"  # Your Telegram user ID, chat ID, or list of IDs

# ======================================================
#                   COLORS + ICONS (IMPORTER)
# ======================================================
C = {
    "INFO":  "\x1b[36m",
    "OK":    "\x1b[32m",
    "WARN":  "\x1b[33m",
    "ERR":   "\x1b[31m",
    "TITLE": "\x1b[35m",
    "STEP":  "\x1b[34m",
    "RESET": "\x1b[0m",
}

if IS_PLAIN:
    C = {k: "" for k in C}
elif IS_MINIMAL:
    C = {k: "\x1b[37m" for k in C}  # white
    C["OK"] = "\x1b[32m"            # green for success
    C["RESET"] = "\x1b[0m"

def now_hms():
    return datetime.now().strftime("%H:%M:%S")

def log(msg, tag="INFO", end="\n"):
    print(f"{C.get(tag,'')}[{now_hms()}] {msg}{C['RESET']}", end=end, flush=True)


# ======================================================
#              TELEGRAM NOTIFICATION
# ======================================================
def send_telegram_notification(project_name: str, video_dir: Path, project_dir: Path):
    """
    Send a Telegram notification with project name, path, and video files list.
    Supports single chat ID or list of chat IDs.
    """
    if not TELEGRAM_BOT_TOKEN or TELEGRAM_BOT_TOKEN == "YOUR_BOT_TOKEN_HERE":
        log("⚠️ Telegram notification skipped (no bot token configured)", "WARN")
        return
    
    if not TELEGRAM_CHAT_ID or TELEGRAM_CHAT_ID == "YOUR_CHAT_ID_HERE":
        log("⚠️ Telegram notification skipped (no chat ID configured)", "WARN")
        return
    
    # Convert single ID to list for uniform handling
    chat_ids = TELEGRAM_CHAT_ID if isinstance(TELEGRAM_CHAT_ID, list) else [TELEGRAM_CHAT_ID]
    
    try:
        # Get list of video files
        video_files = []
        if video_dir.exists():
            video_files = sorted([f.name for f in video_dir.iterdir() if f.is_file()])
        
        # Build message with clean formatting
        message = "✅ *PROJECT COMPLETE*\n\n"
        message += f"📁 *Project:* {project_name}\n"
        message += f"📂 *Path:* `{project_dir}`\n\n"
        
        if video_files:
            message += f"📹 *Files:* {len(video_files)} video(s)\n\n"
            for i, fname in enumerate(video_files, 1):
                message += f"▫️ `{fname}`\n"
        else:
            message += "📭 *No video files found*"
        
        message += "\n\ncc: @salma514"
        
        # Send to all chat IDs
        success_count = 0
        for chat_id in chat_ids:
            try:
                url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
                payload = {
                    "chat_id": chat_id,
                    "text": message,
                    "parse_mode": "Markdown"
                }
                
                response = get_session().post(url, json=payload, timeout=10)
                response.raise_for_status()
                success_count += 1
            except Exception as e:
                log(f"⚠️ Failed to send to chat {chat_id}: {e}", "WARN")
        
        if success_count > 0:
            log(f"📱 Telegram notification sent to {success_count} chat(s)", "OK")
    except Exception as e:
        log(f"⚠️ Failed to send Telegram notification: {e}", "WARN")

# ======================================================
def get_session():
    """
    Returns a requests Session with retry logic for robustness.
    """
    session = requests.Session()
    retry = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["HEAD", "GET", "PUT", "POST", "DELETE"]
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session

# ======================================================
#              PROJECT HANDLING (BY DATE)
# ======================================================
def get_today_project_dir(root: Path, name: str = None) -> Path:
    """
    Create project folder.
    If name is provided, use it.
    Else use today's date format: dd.mm.yyyy
    """
    if name and name.strip():
        # Sanitize name for Windows folder safety
        safe_name = "".join(c for c in name if c not in r'<>:"/\|?*')
        return root / safe_name.strip()
    
    today = datetime.now().strftime("%d.%m.%Y")
    return root / today

def ensure_tree(project_dir: Path):
    project_dir.mkdir(parents=True, exist_ok=True)
    for s in SUBFOLDERS:
        (project_dir / s).mkdir(parents=True, exist_ok=True)

def unique_path(folder: Path, name: str) -> Path:
    base = Path(name).stem
    ext  = Path(name).suffix
    cand = folder / f"{base}{ext}"
    i = 2
    while cand.exists():
        cand = folder / f"{base} ({i}){ext}"
        i += 1
    return cand

# ======================================================
#                 EXTRACT CODES
# ======================================================
VID_CODE_RE = re.compile(r"C(\d+)", re.IGNORECASE)
AUD_CODE_RE = re.compile(r"ZOOM(\d+)", re.IGNORECASE)

def extract_video_code(name):
    m = VID_CODE_RE.search(name)
    if not m:
        return None
    try:
        return int(m.group(1))
    except:
        return None

def extract_audio_code(name):
    m = AUD_CODE_RE.search(name)
    if not m:
        return None
    try:
        return int(m.group(1))
    except:
        return None

# ======================================================
#          WAIT FOR BOTH SOURCES (CAM + CARD)
# ======================================================
# ======================================================
#          WAIT FOR BOTH SOURCES (DYNAMIC SEARCH)
# ======================================================
def get_drives():
    drives = []
    bitmask = ctypes.windll.kernel32.GetLogicalDrives()
    for letter in string.ascii_uppercase:
        if bitmask & 1:
            drives.append(f"{letter}:\\")
        bitmask >>= 1
    return drives

def find_path_on_drives(relative_path: Path):
    """
    Search for `relative_path` on all available drives.
    Returns the first full Path found, or None.
    """
    for drive in get_drives():
        candidate = Path(drive) / relative_path
        if candidate.exists() and candidate.is_dir():
            return candidate
    return None

def wait_for_sources(video_rel: Path, audio_rel: Path, poll_sec: float = 1.0):
    print(f"Waiting for camera ({video_rel}) AND card ({audio_rel}) on ANY drive...")
    
    while True:
        found_video = find_path_on_drives(video_rel)
        found_audio = find_path_on_drives(audio_rel)

        if found_video and found_audio:
            # Check if folders are not empty? (Optional, but good practice per original logic)
            # Original logic: video_src.exists() and any(video_src.iterdir())
            
            # We already checked exists() and is_dir() in find_path_on_drives
            # Let's check contents
            v_has_files = any(found_video.iterdir())
            a_has_files = any(found_audio.iterdir())
            
            if v_has_files and a_has_files:
                print(f"✅ Found Video Source: {found_video}")
                print(f"✅ Found Audio Source: {found_audio}")
                return found_video, found_audio
            else:
                # Found folders but empty? Wait.
                pass
        
        time.sleep(poll_sec)

# ======================================================
#                   DB & DUP GUARD (IMPORTER)
# ======================================================
DB_LOCK = threading.Lock()

def load_db():
    main_db = {"video_codes": [], "audio_codes": [], "video_files": [], "audio_files": []}
    
    def _load(path, target):
        if path.exists():
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
                for key in target:
                    # Sync and Merge items using set to avoid duplicates
                    val = data.get(key, [])
                    if isinstance(val, list):
                        target[key] = list(set(target[key]) | set(val))
            except:
                pass

    # 1. Load from primary
    _load(COPIED_DB_PATH, main_db)
    # 2. Load from fallback (if primary was failing)
    _load(FALLBACK_DB_PATH, main_db)
    
    return main_db

def save_db(db):
    """
    Saves the media history DB with a retry mechanism and fallback location.
    """
    with DB_LOCK:
        tmp = COPIED_DB_PATH.with_suffix(".tmp")
        
        def _write_file(target):
            target.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")

        try:
            # 1. Try Primary Location (F:)
            for attempt in range(5):
                try:
                    if COPIED_DB_PATH.exists():
                        os.chmod(COPIED_DB_PATH, 0o666)
                    
                    _write_file(tmp)
                    if COPIED_DB_PATH.exists():
                        os.remove(COPIED_DB_PATH)
                    shutil.move(str(tmp), str(COPIED_DB_PATH))
                    # If we succeeded on F:, we can optionally keep E: as backup or delete it
                    return 
                except Exception:
                    time.sleep(0.5)
            
            # 2. Fallback to script directory (E:)
            log(f"⚠️ Primary DB save failed, attempting fallback: {FALLBACK_DB_PATH}", "WARN")
            _write_file(FALLBACK_DB_PATH)
            ok("Fallback save successful.")
            
        except Exception as e:
            log(f"⚠️ Critical Save Failure: {e}", "ERR")

def already_copied(db, kind, fname, code):
    base = Path(fname).name.lower().strip()
    
    # Check both code and filename for robustness
    exists_code = code in db.get(f"{kind}_codes", [])
    exists_file = base in db.get(f"{kind}_files", [])
    
    if exists_code or exists_file:
        # If the user says "it still reads them", let's see why it's skipping silently or not
        # log(f"   ⏭  Skipping {fname} (Already in DB)", "INFO")
        return True
    return False

def register_copied(db, kind, fname, code):
    base = Path(fname).name.lower()
    if kind == "video":
        if code not in db["video_codes"]:
            db["video_codes"].append(code)
        if base not in db["video_files"]:
            db["video_files"].append(base)
    else:
        if code not in db["audio_codes"]:
            db["audio_codes"].append(code)
        if base not in db["audio_files"]:
            db["audio_files"].append(base)

# ======================================================
#            COPY WITH LIVE PROGRESS
# ======================================================
def copy_with_progress(src: Path, dst: Path):
    size = src.stat().st_size
    copied = 0
    start = time.time()

    with open(src, "rb") as fsrc, open(dst, "wb") as fdst:
        while True:
            buf = fsrc.read(CHUNK_SIZE)
            if not buf:
                break
            fdst.write(buf)
            copied += len(buf)

            elapsed = max(time.time() - start, 1e-6)
            speed = copied / elapsed
            eta = (size - copied) / speed if speed > 0 else 0
            pct = copied / size if size > 0 else 0
            filled = int(COPY_BAR_WIDTH * pct)
            bar = PROGRESS_FULL * filled + PROGRESS_EMPTY * (COPY_BAR_WIDTH - filled)

            line = (
                f"\r{C['STEP']}   {bar} {pct*100:3.0f}% "
                f"ETA {int(eta//60):02d}:{int(eta%60):02d}{C['RESET']}"
            )
            sys.stdout.write(line)
            sys.stdout.flush()

        fdst.flush()
        os.fsync(fdst.fileno())
    sys.stdout.write("\n")
    sys.stdout.flush()
    return time.time() - start, size

# ======================================================
#                MAIN SCAN LOGIC (IMPORTER)
# ======================================================
def scan_and_copy(src_dir, dst_dir, allow_exts, kind, db, show_progress=True):
    moved = 0
    skipped = 0

    if not src_dir.exists():
        log(f"⚠️ SKIP — Source not found: {src_dir}", "WARN")
        return moved, skipped

    entries = [e for e in src_dir.iterdir() if e.is_file() and e.suffix.lower() in allow_exts]

    log(f"📁 Scanning {kind.upper()} → {src_dir}", "TITLE")
    
    # Pre-scan to count what's actually new
    to_copy = []
    already_in_db = 0
    below_min = 0
    no_code = 0

    for f in entries:
        name = f.name
        if kind == "video":
            code = extract_video_code(name)
            m_code = MIN_VIDEO_CODE
        else:
            code = extract_audio_code(name)
            m_code = MIN_AUDIO_CODE
            
        if code is None:
            no_code += 1
        elif code < m_code:
            below_min += 1
        elif already_copied(db, kind, name, code):
            already_in_db += 1
        else:
            to_copy.append((f, name, code))

    log(f"   Found {len(entries)} total files on source.", "INFO")
    if to_copy:
        log(f"   🚀 NEW FILES TO COPY: {len(to_copy)}", "OK")
    else:
        log(f"   ✅ No new files found to copy.", "OK")
        
    if already_in_db > 0: log(f"   ⏭  Already in DB: {already_in_db}", "INFO")
    if below_min > 0: log(f"   📉 Below min code ({m_code}): {below_min}", "WARN")
    if no_code > 0: log(f"   ❓ No code found in name: {no_code}", "WARN")
    print() # spacing

    total_to_copy = len(to_copy)

    for i, (f, name, code) in enumerate(to_copy, start=1):

        dst = unique_path(dst_dir, name)
        log(f"   [{i}/{len(entries)}] 📦 {name} → {dst.relative_to(dst_dir.parent)}", "STEP")

        try:
            if show_progress:
                dur, _ = copy_with_progress(f, dst)
            else:
                start = time.time()
                with open(f, "rb") as rf, open(dst, "wb") as wf:
                    shutil.copyfileobj(rf, wf, CHUNK_SIZE)
                dur = time.time() - start
            moved += 1
            register_copied(db, kind, name, code)
            # Removed immediate save per file to avoid WinError 5 collisions
            log(f"       ✅ Done in {dur:.2f}s", "OK")
            if not show_progress and kind == "video":
                log(f"       ▶️ VIDEO {i}/{total_to_copy} copied: {name}", "INFO")
        except Exception as e:
            skipped += 1
            log(f"       ❌ ERROR {e}", "ERR")

    skipped += (already_in_db + below_min + no_code)
    if moved > 0:
        save_db(db) # Save once after each directory scan
    return moved, skipped

# ======================================================
#         WAV MERGE + AUPHONIC PIPELINE SECTION
# ======================================================

# =========================
# CLI Styling (pipeline)
# =========================
RESET  = "\033[0m"; BOLD   = "\033[1m"
RED   = "\033[31m"; GREEN = "\033[32m"; YELLOW = "\033[33m"
BLUE  = "\033[34m"; CYAN  = "\033[36m"; MAGENTA = "\033[35m"; BLACK = "\033[30m"
WHITE = "\033[37m"
BGGRE  = "\033[42m"; BGRED = "\033[41m"; BGBLUE = "\033[44m"
BGCYAN = "\033[46m"; BGYEL = "\033[43m"

ICON_START = f"{BGBLUE}{BLACK} ▶ {RESET}"
ICON_OK    = f"{BGGRE}{BLACK} ✓ {RESET}"
ICON_FAIL  = f"{BGRED}{BLACK} ✗ {RESET}"
ICON_WARN  = f"{BGYEL}{BLACK} ! {RESET}"
ICON_INFO  = f"{BGCYAN}{BLACK} i {RESET}"
ICON_TIME  = f"{BGCYAN}{BLACK} ⏱ {RESET}"
ICON_STEP  = f"{MAGENTA}➤{RESET}"
ICON_FILE  = f"{CYAN}📄{RESET}"
ICON_PACK  = f"{YELLOW}🧩{RESET}"

if IS_PLAIN:
    RESET = BOLD = RED = GREEN = YELLOW = BLUE = CYAN = MAGENTA = BLACK = WHITE = ""
    BGGRE = BGRED = BGBLUE = BGCYAN = BGYEL = ""
    ICON_START = "[>]"
    ICON_OK = "[OK]"
    ICON_FAIL = "[X]"
    ICON_WARN = "[!]"
    ICON_INFO = "[i]"
    ICON_TIME = "[time]"
    ICON_STEP = ">"
    ICON_FILE = "[file]"
    ICON_PACK = "[pkg]"
elif IS_MINIMAL:
    # White for everything, highlight start/success only
    RESET = "\033[0m"
    BOLD = ""
    RED = "\033[31m"
    YELLOW = BLUE = MAGENTA = BLACK = ""
    GREEN = "\033[32m"
    CYAN = "\033[36m"
    WHITE = "\033[37m"
    BGGRE = BGRED = BGBLUE = BGCYAN = BGYEL = ""
    ICON_START = f"{CYAN}[>]{RESET}"
    ICON_OK = f"{GREEN}[OK]{RESET}"
    ICON_FAIL = "[X]"
    ICON_WARN = "[!]"
    ICON_INFO = "[i]"
    ICON_TIME = "[time]"
    ICON_STEP = ">"
    ICON_FILE = "[file]"
    ICON_PACK = "[pkg]"

TWO_HOURS_SEC = 7200

# =========================
# Paths for Auphonic config
# =========================
EXCEL_PATH = r"E:\HABASHY\Python Codes\settings.xlsx"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FOLDER = SCRIPT_DIR  # will be set to audio_dir before running the pipeline

# =========================
# Helpers
# =========================
def natural_key(string: str):
    return [int(s) if s.isdigit() else s.lower() for s in re.split(r'(\d+)', string)]

def fmt(sec):
    try:
        sec = int(sec)
    except:
        sec = 0
    return str(timedelta(seconds=sec))

def print_header(title):
    line = "─" * 70
    print(f"\n{BLUE}{line}{RESET}")
    print(f"{ICON_START} {BOLD}{WHITE}{title}{RESET}")
    print(f"{BLUE}{line}{RESET}")

def warn(msg):
    print(f"{ICON_WARN} {YELLOW}{msg}{RESET}")

def fail(msg):
    print(f"{ICON_FAIL} {RED}{msg}{RESET}")

def ok(msg):
    print(f"{ICON_OK} {GREEN}{msg}{RESET}")

def get_duration(path):
    try:
        r = subprocess.run(
            [FFPROBE_PATH, "-v", "error", "-show_entries", "format=duration",
             "-of", "default=noprint_wrappers=1:nokey=1", path],
            capture_output=True, text=True, encoding='utf-8', errors='replace'
        )
        return float(r.stdout.strip())
    except:
        return 0.0

# =========================
# FFmpeg Progress
# =========================
def parse_ffmpeg_time(line):
    if "time=" not in line:
        return None
    m = re.search(r"time=(\d{2}):(\d{2}):(\d{2}\.\d+)", line)
    if not m:
        return None
    h, mnt, s = m.groups()
    return int(h) * 3600 + int(mnt) * 60 + float(s)

def progress_bar(prefix, cur, total):
    pct = 0 if total == 0 else cur / total
    pct = max(0.0, min(1.0, pct))
    filled = int(FFMPEG_BAR_LEN * pct)
    bar = PROGRESS_FULL * filled + PROGRESS_EMPTY * (FFMPEG_BAR_LEN - filled)

    sys.stdout.write(
        f"\r{ICON_STEP} {CYAN}{prefix}{RESET} [{MAGENTA}{bar}{RESET}] "
        f"{YELLOW}{int(pct*100):3d}%{RESET}"
    )
    sys.stdout.flush()

def ffmpeg_run(cmd, total, prefix):
    print(f"{ICON_INFO} {CYAN}Running FFmpeg: {prefix}{RESET}")
    proc = subprocess.Popen(cmd, stderr=subprocess.PIPE, text=True, encoding='utf-8', errors='replace')
    last = 0
    while True:
        line = proc.stderr.readline()
        if not line:
            if proc.poll() is not None:
                break
            continue
        t = parse_ffmpeg_time(line)
        if t is not None:
            last = t
            progress_bar(prefix, last, total)
    print()
    code = proc.wait()
    if code == 0:
        ok(f"FFmpeg step finished: {prefix}")
    else:
        fail(f"FFmpeg exited with code {code} during: {prefix}")
    return code



# =========================
# Naming logic
# =========================
def last3(p):
    nums = re.findall(r"(\d+)", os.path.basename(p))
    return nums[-1][-3:].zfill(3) if nums else "000"

def name_batch(batch):
    a = last3(batch[0])
    b = last3(batch[-1])
    return f"ZOOM-{a}-{b}.mp3"

def name_split(wav):
    a = last3(wav)
    return os.path.join(FOLDER, f"ZOOM-{a}-{a}")

# =========================
# Encode WAV (single file)
# =========================
def encode_single(wav_path, out_path, total_sec):
    cmd = [
        FFMPEG_PATH, "-y",
        "-i", wav_path,
        "-c:a", "libmp3lame", "-b:a", "320k",
        out_path
    ]
    ffmpeg_run(cmd, total_sec, f"encoding {os.path.basename(wav_path)}")
    ok(f"Encoded: {os.path.basename(out_path)}")

def split_long(wav, base, total_sec):
    pattern = f"{base}_part_%02d.mp3"
    cmd = [
        FFMPEG_PATH, "-y", "-i", wav, "-f", "segment",
        "-segment_time", str(TWO_HOURS_SEC),
        "-c:a", "libmp3lame", "-b:a", "320k", pattern
    ]
    ffmpeg_run(cmd, total_sec, "splitting")

    parts = sorted(
        [os.path.join(FOLDER, f) for f in os.listdir(FOLDER)
         if f.startswith(os.path.basename(base)) and f.endswith(".mp3")],
        key=lambda x: natural_key(os.path.basename(x))
    )
    return parts

# =========================
# Read Accounts Excel
# =========================
def load_accounts():
    if not os.path.isfile(EXCEL_PATH):
        fail(f"Excel not found: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb[wb.sheetnames[0]]

    accounts = []
    row = 2

    while True:
        email = sheet[f"A{row}"].value
        api   = sheet[f"B{row}"].value
        preset= sheet[f"C{row}"].value

        if not (email and api and preset):
            break

        accounts.append({
            "email": str(email),
            "api": str(api),
            "preset": str(preset),
            "remaining_minutes": 0.0
        })

        row += 1

    if not accounts:
        fail("No accounts found in Excel.")
        sys.exit(1)

    return accounts

def save_accounts_to_excel(accounts):
    """Overwrites the Excel file with the provided list of active accounts."""
    while True:
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            # Headers
            sheet["A1"] = "Email"
            sheet["B1"] = "API Key"
            sheet["C1"] = "Preset"
            
            for i, acc in enumerate(accounts, start=2):
                sheet[f"A{i}"] = acc["email"]
                sheet[f"B{i}"] = acc["api"]
                sheet[f"C{i}"] = acc["preset"]
                
            wb.save(EXCEL_PATH)
            ok(f"Updated {EXCEL_PATH} (Active accounts: {len(accounts)})")
            break # Success
        except PermissionError:
            fail(f"Permission Denied: Could not save to {EXCEL_PATH}")
            warn("Is the file open in Excel? Please CLOSE IT and press Enter to retry, or Ctrl+C to skip.")
            input("Press Enter to retry...")
        except Exception as e:
            fail(f"Could not update Excel: {e}")
            break

def cleanup_empty_accounts(accounts):
    """Checks for accounts with 0 balance and removes them from Excel."""
    to_keep = [acc for acc in accounts if float(acc.get("remaining_minutes", 0)) > 0.1]
    removed_count = len(accounts) - len(to_keep)
    
    if removed_count > 0:
        warn(f"Detected {removed_count} account(s) with zero balance. Removing from Excel...")
        save_accounts_to_excel(to_keep)
        return to_keep
    return accounts

# =========================
# Auphonic API (Upload Only)
# =========================
def headers(acc):
    return {"Authorization": f"Bearer {acc['api']}"}

def auphonic_create(input_mp3, acc, show_progress=True):
    url = "https://auphonic.com/api/simple/productions.json"

    file_size = os.path.getsize(input_mp3)

    # Use a multipart monitor to show upload progress
    def _make_monitor():
        fields = {
            "preset": acc["preset"],
            "title": os.path.basename(input_mp3),
            "action": "start",
            "input_file": (os.path.basename(input_mp3), open(input_mp3, "rb"), "audio/mpeg")
        }
        encoder = MultipartEncoder(fields=fields)
        start_time = time.time()

        def _cb(monitor):
            pct = monitor.bytes_read / file_size if file_size else 0
            pct = max(0.0, min(1.0, pct))
            elapsed = max(time.time() - start_time, 1e-6)
            speed = monitor.bytes_read / elapsed
            bar_len = 30
            filled = int(bar_len * pct)
            bar = PROGRESS_FULL * filled + PROGRESS_EMPTY * (bar_len - filled)
            eta = (file_size - monitor.bytes_read) / speed if speed > 0 else 0
            sys.stdout.write(
                f"\r{ICON_STEP} Uploading {os.path.basename(input_mp3)} "
                f"[{MAGENTA}{bar}{RESET}] {int(pct*100):3d}%"
            )
            sys.stdout.flush()

        monitor = MultipartEncoderMonitor(encoder, _cb)
        return monitor

    if show_progress:
        monitor = _make_monitor()
        headers_upload = headers(acc)
        headers_upload["Content-Type"] = monitor.content_type
        data = monitor
    else:
        # No monitor, just standard multipart
        m = MultipartEncoder(fields={
            "preset": acc["preset"],
            "title": os.path.basename(input_mp3),
            "action": "start",
            "bitrate": "320",
            "input_file": (os.path.basename(input_mp3), open(input_mp3, "rb"), "audio/mpeg")
        })
        headers_upload = headers(acc)
        headers_upload["Content-Type"] = m.content_type
        data = m

    r = get_session().post(url, headers=headers_upload, data=data)
    if show_progress:
        print()
    r.raise_for_status()
    uuid = r.json()["data"]["uuid"]
    ok(f"Production created: {uuid}")
    return uuid

# =========================
# Auphonic WAIT (No downloading)
# =========================
def auphonic_wait(uuid, acc, audio_duration_seconds, show_progress=True):
    url = f"https://auphonic.com/api/production/{uuid}/status.json"

    start_time = time.time()
    estimated_total = None
    est_start_time = None

    if show_progress:
        print(f"{ICON_INFO} {CYAN}Waiting for Auphonic processing… (UPLOAD ONLY, no download){RESET}")

    while True:
        r = get_session().get(url, headers=headers(acc))
        r.raise_for_status()

        data = r.json()["data"]
        status = data["status"]
        msg    = data["status_string"] or ""
        progress = data.get("progress", 0) or 0  # 0-100 if provided by API

        elapsed = time.time() - start_time


        # Start estimated timer AFTER first real status
        if estimated_total is None and msg not in ["Uploading", "Starting production"]:
            audio_minutes = audio_duration_seconds / 60
            estimated_total = (audio_minutes * 0.05) * 60  # 5% of duration
            est_start_time = time.time()



        short_msg = (msg[:60] + "…" if len(msg) > 60 else msg)

        if show_progress:
            sys.stdout.write(
                f"\r{ICON_TIME} {WHITE}Processing{RESET} "
                f"{CYAN}{short_msg}{RESET} {GREEN}{progress:3.0f}%{RESET}"
            )
            sys.stdout.flush()

        if status == 2:
            if show_progress:
                print()
                ok("Auphonic finished successfully (UPLOAD ONLY, no download)")
            return

        if status == 3:
            if (msg or "").strip().lower() == "done" or progress >= 100:
                if show_progress:
                    print()
                    ok("Auphonic finished (status=3/Done)")
                return
            if show_progress:
                print()
            fail(f"Auphonic error: {msg}")
            # Do not exit the script; continue workflow
            return

        time.sleep(5)

# =========================
# Download Auphonic outputs (for this run only)
# =========================
def download_outputs(uuid, acc, dest_dir: Path, show_progress=True):
    """
    Download all output files for the production we just finished.
    Saved in the same folder as the source mp3.
    """
    url = f"https://auphonic.com/api/production/{uuid}.json"
    try:
        r = get_session().get(url, headers=headers(acc))
        r.raise_for_status()
    except Exception as e:
        fail(f"Fetch production {uuid} failed: {e}")
        return

    data = r.json().get("data", {})
    outs = data.get("output_files", []) or []
    if not outs:
        warn("No output files returned from Auphonic.")
        return

    dest_dir = Path(dest_dir)
    ok(f"Found {len(outs)} output file(s); downloading to {dest_dir}")

    for out in outs:
        dl = out.get("download_url") or ""
        base = out.get("output_basename")
        ending = out.get("ending") or ""

        if dl.startswith("/"):
            dl = "https://auphonic.com" + dl
        if not dl:
            warn(f"Skip file without download URL (uuid {uuid})")
            continue

        filename = None
        if base:
            filename = f"{base}.{ending}" if ending else base

        # If base is missing from Auphonic response, derive filename from URL
        if not filename:
            try:
                parsed = urlparse(dl)
                name = Path(parsed.path).name
                filename = unquote(name.split("?")[0]) if name else None
            except Exception:
                filename = None

        if not filename:
            filename = "output" + (f".{ending}" if ending else "")

        target = dest_dir / filename

        try:
            with get_session().get(dl, headers=headers(acc), stream=True) as resp:
                resp.raise_for_status()
                total = int(resp.headers.get("content-length") or 0)
                downloaded = 0
                start = time.time()
                bar_len = 30

                with open(target, "wb") as f:
                    for chunk in resp.iter_content(8192):
                        if not chunk:
                            continue
                        f.write(chunk)
                        downloaded += len(chunk)

                        if total > 0:
                            pct = downloaded / total
                            filled = int(bar_len * pct)
                            bar = PROGRESS_FULL * filled + PROGRESS_EMPTY * (bar_len - filled)
                            speed = downloaded / max(time.time() - start, 1e-6)
                            eta = (total - downloaded) / speed if speed > 0 else 0
                            if show_progress:
                                sys.stdout.write(
                                    f"\r{ICON_STEP} Downloading {filename} "
                                    f"[{MAGENTA}{bar}{RESET}] {pct*100:3.0f}%"
                                )
                                sys.stdout.flush()
                    if total > 0 and show_progress:
                        sys.stdout.write("\n")
                if show_progress:
                    ok(f"Downloaded: {filename}")
        except Exception as e:
            warn(f"Download failed for {filename}: {e}")

# ======================================================
#  AUPHONIC REAL REMAINING TIME (PLAN A — REAL CREDITS)
# ======================================================
MIN_PER_CREDIT = 60   # Auphonic trial: 1 credit ≈ 60 minutes

def get_account_remaining_minutes(acc):
    """
    Reads remaining credits for an Auphonic account (real-time).
    Returns available minutes as float.
    """
    url = "https://auphonic.com/api/user.json"
    headers = {"Authorization": f"Bearer {acc['api']}"}

    try:
        r = get_session().get(url, headers=headers, timeout=10)
        r.raise_for_status()
        data = r.json().get("data", {})
        credits = float(data.get("credits", 0))
        minutes = credits * MIN_PER_CREDIT
        return minutes
    except Exception as e:
        print(f"{ICON_WARN} Error fetching credits for {acc['email']}: {e}")
        return 0.0

def ensure_min_total_minutes(required_minutes=240):
    """
    Verify total available minutes across all Auphonic accounts before starting.
    Returns False if the total is below required_minutes.
    """
    print(f"{ICON_INFO} {CYAN}Checking Auphonic credits...{RESET}")
    accounts = load_accounts()

    total = 0.0
    
    # Parallel fetch of credits
    with concurrent.futures.ThreadPoolExecutor() as executor:
        # Map each account to the future
        future_to_acc = {executor.submit(get_account_remaining_minutes, acc): acc for acc in accounts}
        
        for future in concurrent.futures.as_completed(future_to_acc):
            acc = future_to_acc[future]
            try:
                rem = future.result()
                acc["remaining_minutes"] = rem
                total += rem
            except Exception as e:
                print(f"{ICON_WARN} Failed to check credits for {acc['email']}: {e}")
                acc["remaining_minutes"] = 0.0

    # NEW: Automatic cleanup of zero-balance accounts
    cleanup_empty_accounts(accounts)

    if total < required_minutes:
        fail(f"Total available minutes {total:.1f} is below required {required_minutes}. Update the Excel sheet then rerun.")
        return False

    ok(f"Total Auphonic minutes available: {total:.1f} (required {required_minutes})")
    return True

# =========================
# PLAN UPLOADS — REAL AUPHONIC MINUTES ONLY
# =========================
def plan_uploads(mp3s, accounts):
    print_header("Auphonic Account Selection (compact)")

    # Fetch real remaining time for each account (silent)
    for acc in accounts:
        acc["remaining_minutes"] = get_account_remaining_minutes(acc)

    uploads = []

    for p in mp3s:
        d_sec = get_duration(p)
        d_min = d_sec / 60 if d_sec > 0 else 0.0

        selected = None

        # Try to find an account that has enough minutes
        for acc in accounts:
            if acc["remaining_minutes"] >= d_min:
                selected = acc
                break

        # Fallback: choose the account that has the highest remaining time
        if selected is None:
            warn(f"SKIPPING {os.path.basename(p)}: No single account has enough credits (Need {d_min:.1f}m).")
            continue

        # Deduct reservation locally
        selected["remaining_minutes"] -= d_min

        uploads.append((p, selected))

        print(f"{ICON_OK} {os.path.basename(p)} → {YELLOW}{selected['email']}{RESET} "
              f"(remaining {selected['remaining_minutes']:.2f} min)")

    return uploads
# =========================
# AUDIO PIPELINE WRAPPER
# =========================
def process_audio_pipeline(audio_dir: Path, db: dict):
    """
    Run WAV merge + MP3 encode + Auphonic upload pipeline
    inside the given audio_dir.
    All outputs/logs stay in the same folder (no moving).
    """
    global FOLDER
    FOLDER = str(audio_dir)

    script_start = time.time()

    print_header("WAV Merge + Auphonic FULL Pipeline")

    accounts = load_accounts()
    print_header("Accounts")
    print(f"{ICON_INFO} Loaded {len(accounts)} account(s)")

    wavs = sorted(
        [os.path.join(FOLDER, f) for f in os.listdir(FOLDER)
         if f.lower().endswith(".wav")],
        key=lambda x: natural_key(os.path.basename(x))
    )


    final_mp3s = []

    if wavs:
        print_header(f"WAV Found ({len(wavs)} file(s))")
    if wavs:
        print_header(f"WAV Found ({len(wavs)} file(s))")
        
        def process_wav(p):
            d = get_duration(p)
            out = os.path.join(FOLDER, f"{Path(p).stem}.mp3")
            
            # Skip if MP3 already exists and looks healthy (size > 1KB)
            if os.path.exists(out) and os.path.getsize(out) > 1024:
                print(f"{ICON_OK} {CYAN}Skipping encoding: {os.path.basename(out)} already exists.{RESET}")
                return [out]

            if d > TWO_HOURS_SEC:
                # Long files split logic
                parts = split_long(p, name_split(p), d)
                return parts
            else:
                encode_single(p, out, d)
                return [out]

        # Parallel encode
        # Adjust max_workers based on CPU cores, e.g., os.cpu_count() or 4
        workers = min(len(wavs), os.cpu_count() or 4)
        print(f"{ICON_INFO} Encoding with {workers} parallel workers...")
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
            results = list(executor.map(process_wav, wavs))
        
        for res in results:
            final_mp3s.extend(res)

    else:
        print_header("No WAV → using existing MP3 (if any)")
        final_mp3s = sorted(
            [os.path.join(FOLDER, f) for f in os.listdir(FOLDER)
             if f.lower().endswith(".mp3")],
            key=lambda x: natural_key(os.path.basename(x))
        )
        if not final_mp3s:
            fail("No audio files found in audio folder. Skipping Auphonic pipeline.")
            return
        print(f"{ICON_INFO} Using {len(final_mp3s)} existing MP3 file(s)")

    # Filter MP3s against local log
    filtered_mp3s = []
    for p in final_mp3s:
        # Additional check: Log file check (as before)
        log_path = os.path.join(FOLDER, "auphonic_account.txt")
        if os.path.exists(log_path):
            with open(log_path, "r", encoding="utf-8") as f:
                if os.path.basename(p) in f.read():
                    print(f"{ICON_OK} {CYAN}Skipping {os.path.basename(p)} (already in log){RESET}")
                    continue
        
        filtered_mp3s.append(p)
    
    to_upload = filtered_mp3s

    if not to_upload:
        ok("All files already processed (found in log).")
        return

    uploads = plan_uploads(to_upload, accounts)

    for p, a in uploads:
        print(f"{ICON_INFO} {WHITE}{os.path.basename(p)}{RESET} "
              f"→ {CYAN}{a['email']}{RESET}")

    print_header("Auphonic Processing (Parallel)")
    log_entries = []
    account_stats = {acc["email"]: {"files": 0, "duration": 0.0} for acc in accounts}

    # Helper for parallel execution
    def process_single_auphonic(item):
        mp3_path, acc = item
        email = acc["email"]
        d = get_duration(mp3_path)
        base = os.path.basename(mp3_path)

        # Thread-safe stats update
        # (Technically dict update is atomic-ish in GIL, but let's be safe if we needed locks. 
        # Here we just update. It's fine for this script complexity.)
        account_stats[email]["files"] += 1
        account_stats[email]["duration"] += d

        print(f"{ICON_START} {BOLD}{WHITE}Start:{RESET} {CYAN}{base}{RESET} → {YELLOW}{email}{RESET}")

        try:
            # 1. Upload (silent)
            uuid = auphonic_create(mp3_path, acc, show_progress=False)
            
            # 2. Wait (silent)
            auphonic_wait(uuid, acc, d, show_progress=False)
            
            # 3. Download (silent)
            download_outputs(uuid, acc, Path(mp3_path).parent.parent / "video", show_progress=False)
            
            print(f"{ICON_OK} {GREEN}Finished:{RESET} {base}")
            return f"{base} → {email}"
        except Exception as e:
            print(f"{ICON_FAIL} {RED}Failed:{RESET} {base} ({e})")
            return None

    # Run in parallel
    # Max workers = number of files, or limit to say 5 to avoid network congestion?
    # Let's use min(len(uploads), 5)
    max_workers = min(len(uploads), 8)
    if max_workers < 1: max_workers = 1

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        results = list(executor.map(process_single_auphonic, uploads))

    # Collect logs
    for res in results:
        if res:
            log_entries.append(res)

    log_path = os.path.join(FOLDER, "auphonic_account.txt")
    with open(log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(log_entries))

    # =========================
    # SUMMARY
    # =========================
    print_header("Summary")

    total_files = len(uploads)
    total_time  = time.time() - script_start
    print(f"{ICON_INFO} {WHITE}Total files processed:{RESET} {CYAN}{total_files}{RESET}")
    print(f"{ICON_INFO} {WHITE}Script runtime:{RESET} {YELLOW}{fmt(total_time)}{RESET}\n")

    ok("ALL DONE (Auphonic Pipeline)!")
    print(f"{ICON_INFO} {CYAN}Log saved in:{RESET} {WHITE}{FOLDER}{RESET}")
    print(f"{ICON_INFO} {CYAN}Outputs saved in:{RESET} {WHITE}{Path(FOLDER).parent / 'video'}{RESET}")
    print(f"{ICON_FILE} {WHITE}{log_path}{RESET}")

# ======================================================
#                  MAIN (FULL WORKFLOW)
# ======================================================
def get_latest_project_dir(root: Path):
    if not root.exists():
        return None
    dirs = [d for d in root.iterdir() if d.is_dir()]
    if not dirs:
        return None
    # Sort by creation time, descending
    dirs.sort(key=lambda d: os.path.getctime(d), reverse=True)
    return dirs[0]

def main():
    # Print script last modification time
    try:
        mod_time = os.path.getmtime(__file__)
        dt_mod = datetime.fromtimestamp(mod_time).strftime("%Y-%m-%d %H:%M:%S")
        print(f"\n🕒 Script Last Modified: {dt_mod}")
    except Exception:
        pass

    if not BASE_ROOT.exists():
        log(f"❌ BASE ROOT does not exist: {BASE_ROOT}", "ERR")
        sys.exit(1)

    # Early check on Auphonic credit before any copying (BYPASSED)
    # if not ensure_min_total_minutes(240):
    #     log("⚠️ Insufficient Auphonic credits, but continuing as requested (Import + MP3 conversion)...", "WARN")

    # Wait until both camera and card are ready (not just one)
    # This now scans all drives dynamically and returns the actual paths
    real_video_src, real_audio_src = wait_for_sources(VIDEO_SRC_REL, AUDIO_SRC_REL)

    log("⚡ Auto Project Importer — LIVE", "TITLE")

    # ========================
    # GET PROJECT NAME (Arabic Support)
    # ========================
    
    # Check for latest project
    latest_project = get_latest_project_dir(BASE_ROOT)
    use_latest = False
    project_name = ""

    if latest_project:
        print(f"\n{C['TITLE']}{'='*70}{C['RESET']}")
        print(f"{C['TITLE']}  PROJECT SELECTION{C['RESET']}")
        print(f"{C['TITLE']}{'='*70}{C['RESET']}")
        print(f"1. Create New Project (default)")
        print(f"2. Use Last Project: {latest_project.name}")
        
        choice = input(f"{C['STEP']}Choose (1/2): {C['RESET']}").strip()
        if choice == "2":
            use_latest = True
            project = latest_project
            project_name = project.name
            log(f"✅ Using existing project: {project_name}", "OK")

    if not use_latest:
        print(f"\n{C['TITLE']}{'='*70}{C['RESET']}")
        print(f"{C['TITLE']}  PROJECT NAME{C['RESET']}")
        print(f"{C['TITLE']}{'='*70}{C['RESET']}")
        print(f"{C['STEP']}Enter project name (or press Enter to use date): {C['RESET']}", end="", flush=True)
        
        try:
            # Read input with UTF-8 encoding for Arabic support
            project_name = input().strip()
        except (UnicodeDecodeError, KeyboardInterrupt):
            project_name = ""
        
        if project_name:
            # Apply RTL formatting for Arabic text
            display_name = get_display(project_name) if HAS_BIDI else project_name
            log(f"✅ Project: {display_name}", "OK")
        else:
            log(f"📅 Using date-based naming", "INFO")
        
        project = get_today_project_dir(BASE_ROOT, project_name if project_name else None)
    ensure_tree(project)
    log(f"📁 PROJECT (by date) → {project}", "INFO")
    for s in SUBFOLDERS:
        log(f"   ensured: {project / s}", "INFO")

    audio_dir = project / "audio"
    video_dir = project / "video"

    db = load_db()
    log(f"🗂 Using DB → {COPIED_DB_PATH}", "INFO")

    # --- SHOW DATABASE SUMMARY ---
    last_v = max(db.get("video_codes", [0])) if db.get("video_codes") else "None"
    last_a = max(db.get("audio_codes", [0])) if db.get("audio_codes") else "None"
    print(f"\n{C['INFO']}📊 DATABASE SUMMARY:{C['RESET']}")
    print(f"   - Last Video Code: {C['OK']}{last_v}{C['RESET']}")
    print(f"   - Last Audio Code: {C['OK']}{last_a}{C['RESET']}")
    print(f"   - Total Files in DB: {len(db.get('video_files', [])) + len(db.get('audio_files', []))}\n")

    # ========================
    # TIMING STATS
    # ========================
    stats = {
        "start_total": time.time(),
        "audio_import": 0.0,
        "video_import": 0.0,
        "auphonic": 0.0
    }

    # ========================
    # 1) AUDIO IMPORT FIRST
    # ========================
    log("\n🎧 AUDIO IMPORT", "TITLE")
    t0 = time.time()
    am, as_ = scan_and_copy(real_audio_src, audio_dir, AUDIO_EXTS, "audio", db)
    stats["audio_import"] = time.time() - t0
    log(f"   ✅ moved:{am} , ⏭ skipped:{as_}", "OK")

    # ========================
    # 2) VIDEO IMPORT + AUPHONIC IN PARALLEL
    # ========================
    video_result = {"moved": 0, "skipped": 0}
    pipeline_error = {"err": None}

    def _video_job():
        # Disable progress bar for video copy to reduce overlap with audio pipeline
        t0 = time.time()
        vm, vs = scan_and_copy(real_video_src, video_dir, VIDEO_EXTS, "video", db, show_progress=False)
        stats["video_import"] = time.time() - t0
        video_result["moved"] = vm
        video_result["skipped"] = vs

    def _pipeline_job():
        log("\n🔄 Starting Auphonic Pipeline…", "INFO")
        try:
            t0 = time.time()
            process_audio_pipeline(audio_dir, db)
            stats["auphonic"] = time.time() - t0
            log("🎧 Auphonic pipeline finished — continuing…", "OK")
        except Exception as e:
            pipeline_error["err"] = e
            fail(f"Auphonic pipeline exception: {e}")
            log("⚠️ Continuing workflow despite the pipeline error…", "WARN")

    # Start pipeline first so it shows in logs, then start video copy in parallel
    pipeline_thread = threading.Thread(target=_pipeline_job, name="AuphonicPipeline", daemon=True)
    video_thread = threading.Thread(target=_video_job, name="VideoImport", daemon=True)

    pipeline_thread.start()
    log("\n🎥 Starting VIDEO import in parallel with Auphonic…", "INFO")
    video_thread.start()

    pipeline_thread.join()
    video_thread.join()

    # ========================
    # 3) VIDEO IMPORT SUMMARY
    # ========================
    vm, vs = video_result["moved"], video_result["skipped"]
    log("\n🎥 VIDEO IMPORT (parallel)", "TITLE")
    log(f"   ✅ moved:{vm} , ⏭ skipped:{vs}", "OK")

    save_db(db)
    log("💾 DB updated", "INFO")

    log("\n✅ ALL DONE.", "OK")

    # ========================
    # FINAL TIMING SUMMARY
    # ========================
    total_duration = time.time() - stats["start_total"]
    
    print_header("⏱ EXECUTION TIME SUMMARY")
    print(f"{ICON_INFO} {WHITE}Audio Import:{RESET}    {CYAN}{fmt(stats['audio_import'])}{RESET}")
    print(f"{ICON_INFO} {WHITE}Video Import:{RESET}    {CYAN}{fmt(stats['video_import'])}{RESET}")
    print(f"{ICON_INFO} {WHITE}Auphonic Pipe:{RESET}   {CYAN}{fmt(stats['auphonic'])}{RESET}")
    print(f"{ICON_INFO} {WHITE}Total Runtime:{RESET}   {YELLOW}{fmt(total_duration)}{RESET}")
    print(f"{BLUE}──────────────────────────────────────────────────────────────────────{RESET}")

    # ========================
    # SEND TELEGRAM NOTIFICATION
    # ========================
    send_telegram_notification(project_name if project_name else datetime.now().strftime("%d.%m.%Y"), video_dir, project)




if __name__ == "__main__":
    main()
    input("Press Enter to exit...")
