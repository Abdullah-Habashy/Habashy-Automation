#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Auphonic Standalone Pipeline
---------------------------------------------------------
This script takes a folder full of audio files or a single audio file as input.
It runs the exact parallel conversion, upload, and download pipeline
as in the main Auto Project Importer script.
"""

import os
import sys
import time
import re
import subprocess
import shutil
import concurrent.futures
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl
import colorama
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from urllib.parse import urlparse, unquote
from requests_toolbelt.multipart.encoder import MultipartEncoder, MultipartEncoderMonitor

colorama.just_fix_windows_console()

# ======================================================
#                 CONFIG & PATHS
# ======================================================
def find_tool(name, fallback):
    local_bin = Path(__file__).parent / "bin" / f"{name}.exe"
    if local_bin.exists(): return str(local_bin)
    system_path = shutil.which(name)
    if system_path: return system_path
    return fallback

BIN_FALLBACK = r"E:\HABASHY\ffmpeg-2025-10-21-git-535d4047d3-full_build\ffmpeg-2025-10-21-git-535d4047d3-full_build\bin"
FFMPEG_PATH = find_tool("ffmpeg", os.path.join(BIN_FALLBACK, "ffmpeg.exe"))
FFPROBE_PATH = find_tool("ffprobe", os.path.join(BIN_FALLBACK, "ffprobe.exe"))

EXCEL_PATH = r"E:\HABASHY\Python Codes\settings.xlsx"

TWO_HOURS_SEC = 7200
CHUNK_SIZE = 4 * 1024 * 1024

# ======================================================
#                 UI STYLING
# ======================================================
CLI_MODE = "minimal"
IS_PLAIN = CLI_MODE == "plain"
IS_MINIMAL = CLI_MODE == "minimal"
IS_FULL = CLI_MODE == "full"

RESET  = "\033[0m"; BOLD   = "\033[1m"
RED   = "\033[31m"; GREEN = "\033[32m"; YELLOW = "\033[33m"
BLUE  = "\033[34m"; CYAN  = "\033[36m"; MAGENTA = "\033[35m"; BLACK = "\033[30m"
WHITE = "\033[37m"
BGGRE  = "\033[42m"; BGRED = "\033[41m"; BGBLUE = "\033[44m"
BGCYAN = "\033[46m"; BGYEL = "\033[43m"

ICON_START = f"{BGBLUE}{BLACK} ▶ {RESET}"
ICON_OK    = f"{BGGRE}{BLACK} ✓ {RESET}"
ICON_FAIL  = f"{BGRED}{BLACK} ✗ {RESET}"
ICON_WARN  = f"{BGYEL}{BLACK} ! {RESET}"
ICON_INFO  = f"{BGCYAN}{BLACK} i {RESET}"
ICON_TIME  = f"{BGCYAN}{BLACK} ⏱ {RESET}"
ICON_STEP  = f"{MAGENTA}➤{RESET}"
ICON_FILE  = f"{CYAN}📄{RESET}"
ICON_PACK  = f"{YELLOW}🧩{RESET}"

if IS_PLAIN:
    RESET = BOLD = RED = GREEN = YELLOW = BLUE = CYAN = MAGENTA = BLACK = WHITE = ""
    BGGRE = BGRED = BGBLUE = BGCYAN = BGYEL = ""
    ICON_START = "[>]"; ICON_OK = "[OK]"; ICON_FAIL = "[X]"; ICON_WARN = "[!]"
    ICON_INFO = "[i]"; ICON_TIME = "[time]"; ICON_STEP = ">"; ICON_FILE = "[file]"; ICON_PACK = "[pkg]"
elif IS_MINIMAL:
    RESET = "\033[0m"; BOLD = ""
    RED = "\033[31m"; YELLOW = BLUE = MAGENTA = BLACK = ""
    GREEN = "\033[32m"; CYAN = "\033[36m"; WHITE = "\033[37m"
    BGGRE = BGRED = BGBLUE = BGCYAN = BGYEL = ""
    ICON_START = f"{CYAN}[>]{RESET}"; ICON_OK = f"{GREEN}[OK]{RESET}"
    ICON_FAIL = "[X]"; ICON_WARN = "[!]"; ICON_INFO = "[i]"
    ICON_TIME = "[time]"; ICON_STEP = ">"; ICON_FILE = "[file]"; ICON_PACK = "[pkg]"

PROGRESS_FULL = "#" if (IS_PLAIN or IS_MINIMAL) else "█"
PROGRESS_EMPTY = "." if (IS_PLAIN or IS_MINIMAL) else "░"
FFMPEG_BAR_LEN = 32 if (IS_PLAIN or IS_MINIMAL) else 40

# ======================================================
#                 HELPERS
# ======================================================
def get_session():
    if not hasattr(get_session, "_session"):
        session = requests.Session()
        retry = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504], allowed_methods=["HEAD", "GET", "PUT", "POST", "DELETE"])
        adapter = HTTPAdapter(max_retries=retry)
        session.mount("https://", adapter)
        session.mount("http://", adapter)
        get_session._session = session
    return get_session._session

def natural_key(string: str):
    return [int(s) if s.isdigit() else s.lower() for s in re.split(r'(\d+)', string)]

def fmt(sec):
    try: sec = int(sec)
    except: sec = 0
    return str(timedelta(seconds=sec))

def print_header(title):
    line = "─" * 70
    print(f"\n{BLUE}{line}{RESET}")
    print(f"{ICON_START} {BOLD}{WHITE}{title}{RESET}")
    print(f"{BLUE}{line}{RESET}")

def warn(msg): print(f"{ICON_WARN} {YELLOW}{msg}{RESET}")
def fail(msg): print(f"{ICON_FAIL} {RED}{msg}{RESET}")
def ok(msg):   print(f"{ICON_OK} {GREEN}{msg}{RESET}")

def get_duration(path):
    try:
        r = subprocess.run(
            [FFPROBE_PATH, "-v", "error", "-show_entries", "format=duration",
             "-of", "default=noprint_wrappers=1:nokey=1", path],
            capture_output=True, text=True, encoding='utf-8', errors='replace'
        )
        return float(r.stdout.strip())
    except:
        return 0.0

def parse_ffmpeg_time(line):
    if "time=" not in line: return None
    m = re.search(r"time=(\d{2}):(\d{2}):(\d{2}\.\d+)", line)
    if not m: return None
    h, mnt, s = m.groups()
    return int(h) * 3600 + int(mnt) * 60 + float(s)

def progress_bar(prefix, cur, total):
    pct = 0 if total == 0 else cur / total
    pct = max(0.0, min(1.0, pct))
    filled = int(FFMPEG_BAR_LEN * pct)
    bar = PROGRESS_FULL * filled + PROGRESS_EMPTY * (FFMPEG_BAR_LEN - filled)
    sys.stdout.write(f"\r{ICON_STEP} {CYAN}{prefix}{RESET} [{MAGENTA}{bar}{RESET}] {YELLOW}{int(pct*100):3d}%{RESET}")
    sys.stdout.flush()

def ffmpeg_run(cmd, total, prefix):
    print(f"{ICON_INFO} {CYAN}Running FFmpeg: {prefix}{RESET}")
    proc = subprocess.Popen(cmd, stderr=subprocess.PIPE, text=True, encoding='utf-8', errors='replace')
    last = 0
    while True:
        line = proc.stderr.readline()
        if not line:
            if proc.poll() is not None: break
            continue
        t = parse_ffmpeg_time(line)
        if t is not None:
            last = t
            progress_bar(prefix, last, total)
    print()
    code = proc.wait()
    if code == 0: ok(f"FFmpeg step finished: {prefix}")
    else: fail(f"FFmpeg exited with code {code} during: {prefix}")
    return code

def last3(p):
    nums = re.findall(r"(\d+)", os.path.basename(p))
    return nums[-1][-3:].zfill(3) if nums else "000"

def name_split(wav, folder):
    a = last3(wav)
    return os.path.join(folder, f"ZOOM-{a}-{a}")

def encode_single(wav_path, out_path, total_sec):
    cmd = [FFMPEG_PATH, "-y", "-i", wav_path, "-c:a", "libmp3lame", "-b:a", "320k", out_path]
    ffmpeg_run(cmd, total_sec, f"encoding {os.path.basename(wav_path)}")
    ok(f"Encoded: {os.path.basename(out_path)}")

def split_long(wav, base, total_sec):
    pattern = f"{base}_part_%02d.mp3"
    cmd = [FFMPEG_PATH, "-y", "-i", wav, "-f", "segment", "-segment_time", str(TWO_HOURS_SEC), "-c:a", "libmp3lame", "-b:a", "320k", pattern]
    ffmpeg_run(cmd, total_sec, "splitting")
    folder = os.path.dirname(wav)
    parts = sorted([os.path.join(folder, f) for f in os.listdir(folder) if f.startswith(os.path.basename(base)) and f.endswith(".mp3")], key=lambda x: natural_key(os.path.basename(x)))
    return parts

# ======================================================
#                 ACCOUNTS & API
# ======================================================
def load_accounts():
    if not os.path.isfile(EXCEL_PATH):
        fail(f"Excel not found: {EXCEL_PATH}"); sys.exit(1)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb[wb.sheetnames[0]]
    accounts = []
    row = 2
    while True:
        email = sheet[f"A{row}"].value
        api   = sheet[f"B{row}"].value
        preset= sheet[f"C{row}"].value
        if not (email and api and preset): break
        accounts.append({"email": str(email), "api": str(api), "preset": str(preset), "remaining_minutes": 0.0})
        row += 1
    if not accounts:
        fail("No accounts found in Excel."); sys.exit(1)
    return accounts

def save_accounts_to_excel(accounts):
    while True:
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet["A1"] = "Email"; sheet["B1"] = "API Key"; sheet["C1"] = "Preset"
            for i, acc in enumerate(accounts, start=2):
                sheet[f"A{i}"] = acc["email"]
                sheet[f"B{i}"] = acc["api"]
                sheet[f"C{i}"] = acc["preset"]
            wb.save(EXCEL_PATH)
            ok(f"Updated {EXCEL_PATH} (Active accounts: {len(accounts)})")
            break
        except PermissionError:
            warn("Is the file open in Excel? Please CLOSE IT and press Enter to retry, or Ctrl+C to skip.")
            input("Press Enter to retry...")
        except Exception as e:
            fail(f"Could not update Excel: {e}")
            break

def cleanup_empty_accounts(accounts):
    to_keep = [acc for acc in accounts if float(acc.get("remaining_minutes", 0)) > 0.1]
    removed_count = len(accounts) - len(to_keep)
    if removed_count > 0:
        warn(f"Detected {removed_count} account(s) with zero balance. Removing from Excel...")
        save_accounts_to_excel(to_keep)
        return to_keep
    return accounts

def get_account_remaining_minutes(acc):
    url = "https://auphonic.com/api/user.json"
    headers = {"Authorization": f"Bearer {acc['api']}"}
    try:
        r = get_session().get(url, headers=headers, timeout=10)
        r.raise_for_status()
        return float(r.json().get("data", {}).get("credits", 0)) * 60
    except Exception as e:
        print(f"{ICON_WARN} Error fetching credits for {acc['email']}: {e}")
        return 0.0

def headers(acc): return {"Authorization": f"Bearer {acc['api']}"}

def auphonic_create(input_mp3, acc, show_progress=True):
    url = "https://auphonic.com/api/simple/productions.json"
    file_size = os.path.getsize(input_mp3)

    def _make_monitor():
        fields = {"preset": acc["preset"], "title": os.path.basename(input_mp3), "action": "start", "input_file": (os.path.basename(input_mp3), open(input_mp3, "rb"), "audio/mpeg")}
        encoder = MultipartEncoder(fields=fields)
        start_time = time.time()
        def _cb(monitor):
            pct = monitor.bytes_read / file_size if file_size else 0
            pct = max(0.0, min(1.0, pct))
            elapsed = max(time.time() - start_time, 1e-6)
            bar_len = 30
            filled = int(bar_len * pct)
            bar = PROGRESS_FULL * filled + PROGRESS_EMPTY * (bar_len - filled)
            sys.stdout.write(f"\r{ICON_STEP} Uploading {os.path.basename(input_mp3)} [{MAGENTA}{bar}{RESET}] {int(pct*100):3d}%")
            sys.stdout.flush()
        return MultipartEncoderMonitor(encoder, _cb)

    if show_progress:
        monitor = _make_monitor()
        headers_upload = headers(acc)
        headers_upload["Content-Type"] = monitor.content_type
        data = monitor
    else:
        m = MultipartEncoder(fields={"preset": acc["preset"], "title": os.path.basename(input_mp3), "action": "start", "bitrate": "320", "input_file": (os.path.basename(input_mp3), open(input_mp3, "rb"), "audio/mpeg")})
        headers_upload = headers(acc)
        headers_upload["Content-Type"] = m.content_type
        data = m

    r = get_session().post(url, headers=headers_upload, data=data)
    if show_progress: print()
    r.raise_for_status()
    uuid = r.json()["data"]["uuid"]
    ok(f"Production created: {uuid}")
    return uuid

def auphonic_wait(uuid, acc, audio_duration_seconds, show_progress=True):
    url = f"https://auphonic.com/api/production/{uuid}/status.json"
    start_time = time.time()
    
    if show_progress:
        print(f"{ICON_INFO} {CYAN}Waiting for Auphonic processing… (UPLOAD ONLY){RESET}")

    while True:
        r = get_session().get(url, headers=headers(acc))
        r.raise_for_status()

        data = r.json()["data"]
        status = data["status"]
        msg    = data["status_string"] or ""
        progress = data.get("progress", 0) or 0

        short_msg = (msg[:60] + "…" if len(msg) > 60 else msg)

        if show_progress:
            sys.stdout.write(f"\r{ICON_TIME} {WHITE}Processing{RESET} {CYAN}{short_msg}{RESET} {GREEN}{progress:3.0f}%{RESET}")
            sys.stdout.flush()

        if status == 2:
            if show_progress:
                print(); ok("Auphonic finished successfully (UPLOAD ONLY)")
            return
        if status == 3:
            if (msg or "").strip().lower() == "done" or progress >= 100:
                if show_progress:
                    print(); ok("Auphonic finished (status=3/Done)")
                return
            if show_progress: print()
            fail(f"Auphonic error: {msg}")
            return
        time.sleep(5)

def download_outputs(uuid, acc, dest_dir: Path, show_progress=True):
    url = f"https://auphonic.com/api/production/{uuid}.json"
    try:
        r = get_session().get(url, headers=headers(acc))
        r.raise_for_status()
    except Exception as e:
        fail(f"Fetch production {uuid} failed: {e}")
        return

    data = r.json().get("data", {})
    outs = data.get("output_files", []) or []
    if not outs:
        warn("No output files returned from Auphonic.")
        return

    dest_dir = Path(dest_dir)
    ok(f"Found {len(outs)} output file(s); downloading to {dest_dir}")

    for out in outs:
        dl = out.get("download_url") or ""
        base = out.get("output_basename")
        ending = out.get("ending") or ""

        if dl.startswith("/"): dl = "https://auphonic.com" + dl
        if not dl: continue

        filename = None
        if base: filename = f"{base}.{ending}" if ending else base

        if not filename:
            try:
                parsed = urlparse(dl)
                name = Path(parsed.path).name
                filename = unquote(name.split("?")[0]) if name else None
            except Exception:
                filename = None

        if not filename:
            filename = "output" + (f".{ending}" if ending else "")

        target = dest_dir / filename

        try:
            with get_session().get(dl, headers=headers(acc), stream=True) as resp:
                resp.raise_for_status()
                total = int(resp.headers.get("content-length") or 0)
                downloaded = 0
                bar_len = 30

                with open(target, "wb") as f:
                    for chunk in resp.iter_content(8192):
                        if not chunk: continue
                        f.write(chunk)
                        downloaded += len(chunk)

                        if total > 0 and show_progress:
                            pct = downloaded / total
                            filled = int(bar_len * pct)
                            bar = PROGRESS_FULL * filled + PROGRESS_EMPTY * (bar_len - filled)
                            sys.stdout.write(f"\r{ICON_STEP} Downloading {filename} [{MAGENTA}{bar}{RESET}] {pct*100:3.0f}%")
                            sys.stdout.flush()
                    if total > 0 and show_progress: sys.stdout.write("\n")
                if show_progress: ok(f"Downloaded: {filename}")
        except Exception as e:
            warn(f"Download failed for {filename}: {e}")

def plan_uploads(mp3s, accounts):
    print_header("Auphonic Account Selection (compact)")
    for acc in accounts:
        acc["remaining_minutes"] = get_account_remaining_minutes(acc)

    uploads = []
    for p in mp3s:
        d_min = get_duration(p) / 60 if get_duration(p) > 0 else 0.0
        selected = None
        for acc in accounts:
            if acc["remaining_minutes"] >= d_min:
                selected = acc
                break

        if selected is None:
            warn(f"SKIPPING {os.path.basename(p)}: No single account has enough credits (Need {d_min:.1f}m).")
            continue

        selected["remaining_minutes"] -= d_min
        uploads.append((p, selected))
        print(f"{ICON_OK} {os.path.basename(p)} → {YELLOW}{selected['email']}{RESET} (remaining {selected['remaining_minutes']:.2f} min)")

    return uploads

# ======================================================
#                 MAIN PIPELINE
# ======================================================
def process_audio_pipeline(files_list, parent_folder):
    if not files_list:
        fail("No audio files provided.")
        return

    log_path = os.path.join(parent_folder, "auphonic_account.txt")
    script_start = time.time()

    print_header("WAV Merge + Auphonic FULL Pipeline")

    accounts = load_accounts()
    print_header("Accounts")
    print(f"{ICON_INFO} Loaded {len(accounts)} account(s)")

    # 1. Gather WAVs and MP3s
    wavs = sorted([p for p in files_list if p.lower().endswith(".wav")], key=lambda x: natural_key(os.path.basename(x)))
    pre_mp3s = sorted([p for p in files_list if p.lower().endswith(".mp3")], key=lambda x: natural_key(os.path.basename(x)))

    final_mp3s = list(pre_mp3s)

    if wavs:
        print_header(f"WAV Found ({len(wavs)} file(s))")
        def process_wav(p):
            d = get_duration(p)
            out = os.path.join(parent_folder, f"{Path(p).stem}.mp3")
            if os.path.exists(out) and os.path.getsize(out) > 1024:
                print(f"{ICON_OK} {CYAN}Skipping encoding: {os.path.basename(out)} already exists.{RESET}")
                return [out]

            if d > TWO_HOURS_SEC:
                parts = split_long(p, name_split(p, parent_folder), d)
                return parts
            else:
                encode_single(p, out, d)
                return [out]

        workers = min(len(wavs), os.cpu_count() or 4)
        print(f"{ICON_INFO} Encoding with {workers} parallel workers...")
        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
            results = list(executor.map(process_wav, wavs))
        
        for res in results:
            final_mp3s.extend(res)
    else:
        print_header("No WAVs → using existing MP3s (if any)")
        if not final_mp3s:
            fail("No audio files found. Skipping Auphonic pipeline.")
            return
        print(f"{ICON_INFO} Using {len(final_mp3s)} existing MP3 file(s)")

    # 2. Filter MP3s against local log
    filtered_mp3s = []
    # Using a set to prevent duplicates if a wav and mp3 resolve to the same encoded mp3 path
    seen = set()
    for p in final_mp3s:
        if p in seen: continue
        seen.add(p)
        
        if os.path.exists(log_path):
            with open(log_path, "r", encoding="utf-8") as f:
                if os.path.basename(p) in f.read():
                    print(f"{ICON_OK} {CYAN}Skipping {os.path.basename(p)} (already in log){RESET}")
                    continue
        filtered_mp3s.append(p)
    
    to_upload = filtered_mp3s

    if not to_upload:
        ok("All files already processed (found in log).")
        return

    uploads = plan_uploads(to_upload, accounts)
    for p, a in uploads:
        print(f"{ICON_INFO} {WHITE}{os.path.basename(p)}{RESET} → {CYAN}{a['email']}{RESET}")

    # 3. Parallel Auphonic Processing
    print_header("Auphonic Processing (Parallel)")
    log_entries = []

    def process_single_auphonic(item):
        mp3_path, acc = item
        email = acc["email"]
        d = get_duration(mp3_path)
        base = os.path.basename(mp3_path)

        print(f"{ICON_START} {BOLD}{WHITE}Start:{RESET} {CYAN}{base}{RESET} → {YELLOW}{email}{RESET}")

        try:
            # 1. Upload (silent)
            uuid = auphonic_create(mp3_path, acc, show_progress=False)
            # 2. Wait (silent)
            auphonic_wait(uuid, acc, d, show_progress=False)
            # 3. Download (silent) to the same directory as the MP3
            download_outputs(uuid, acc, Path(mp3_path).parent, show_progress=False)
            
            print(f"{ICON_OK} {GREEN}Finished:{RESET} {base}")
            return f"{base} → {email}"
        except Exception as e:
            print(f"{ICON_FAIL} {RED}Failed:{RESET} {base} ({e})")
            return None

    max_workers = min(len(uploads), 8)
    if max_workers < 1: max_workers = 1

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        results = list(executor.map(process_single_auphonic, uploads))

    for res in results:
        if res:
            log_entries.append(res)

    with open(log_path, "a", encoding="utf-8") as f:
        f.write("\n" + "\n".join(log_entries))

    # 4. Summary & Cleanup
    cleanup_empty_accounts(accounts)

    print_header("Summary")
    total_files = len(uploads)
    total_time  = time.time() - script_start
    print(f"{ICON_INFO} {WHITE}Total files processed:{RESET} {CYAN}{total_files}{RESET}")
    print(f"{ICON_INFO} {WHITE}Script runtime:{RESET} {YELLOW}{fmt(total_time)}{RESET}\n")

    ok("ALL DONE (Auphonic Pipeline)!")
    print(f"{ICON_INFO} {CYAN}Log saved in:{RESET} {WHITE}{parent_folder}{RESET}")

def main():
    print_header("Auphonic Standalone Processor")
    print(f"{ICON_INFO} Enter a folder path or a direct file path:")
    
    choice = input(f"{ICON_STEP} Path: ").strip().strip('"').strip("'")
    if not choice:
        fail("No path provided.")
        return

    target_path = Path(choice)
    if not target_path.exists():
        fail("The provided path does not exist.")
        return

    files_list = []
    if target_path.is_file():
        if target_path.suffix.lower() not in [".wav", ".mp3"]:
            fail("The provided file is not a WAV or MP3.")
            return
        files_list = [str(target_path)]
        parent_folder = str(target_path.parent)
    elif target_path.is_dir():
        files_list = [str(f) for f in target_path.iterdir() if f.is_file() and f.suffix.lower() in [".wav", ".mp3"]]
        parent_folder = str(target_path)
    
    if not files_list:
        fail("No valid audio files (WAV/MP3) found.")
        return

    process_audio_pipeline(files_list, parent_folder)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        warn("\nTerminated by user.")
    except Exception as e:
        fail(f"\nAn error occurred: {e}")
    finally:
        input("\nPress Enter to exit...")
